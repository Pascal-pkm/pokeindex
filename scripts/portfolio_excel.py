# -*- coding: utf-8 -*-
"""Excel-Export der Portfolio-Analyse.

Konventionen: Arial durchgehend, echte Formeln statt vorberechneter Werte
(die Mappe rechnet nach, wenn Preise oder Gebühren geändert werden), blaue
Schrift für Eingaben, schwarze für Formeln, gelbe Füllung für Annahmen.
Jede Annahme steht in einer eigenen benannten Zelle und wird referenziert.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

FONT = "Arial"
INK = Font(name=FONT, size=10)
INK_B = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True)
INPUT = Font(name=FONT, size=10, color="0000FF")
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(name=FONT, size=10, bold=True, color="FFFFFF")
ASSUM_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(bottom=THIN)

EUR = '#,##0.00 "€";(#,##0.00 "€");-'
PCT = '0.0%;(0.0%);-'
NUM = '#,##0;(#,##0);-'


def _head(ws, row: int, labels) -> None:
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _widths(ws, widths) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_excel(path: str, result: dict) -> None:
    wb = Workbook()

    # ------------------------------------------------------------ Annahmen
    ws = wb.active
    ws.title = "Annahmen"
    ws["A1"] = "Portfolio-Analyse – Annahmen und Legende"
    ws["A1"].font = TITLE
    # Jede Annahme erhält einen DEFINIERTEN NAMEN. Formeln referenzieren nur
    # diese Namen, nie Zeilennummern: eine zusätzliche Zeile im Kopf hätte
    # sonst stillschweigend auf die falsche Zelle gezeigt (genau dieser Fehler
    # ist im ersten Testlauf aufgetreten – die Mappe rechnete fehlerfrei, aber
    # mit einem Datumswert als Gebührensatz).
    rows = [
        ("Stand der Bewertung", result["summary"]["stand"], None, None),
        ("Erstellt (UTC)", result["built"], None, None),
        ("Methodikversion", result["method_version"], None, None),
        ("Verkaufsgebühr Cardmarket", 0.06,
         "5 % Provision + ~1 % Zahlungsabwicklung (Stand 07/2026)", "gebuehr_cm"),
        ("Verkaufsgebühr eBay.de (privat)", 0.11,
         "11 % bis 1.990 €, zzgl. 0,35 € Fixgebühr je Bestellung", "gebuehr_ebay"),
        ("Verkaufsgebühr Amazon.de", 0.15,
         "15 % Kategorieprovision (Spielwaren)", "gebuehr_amazon"),
        ("Risikofreier Zins p. a.", 0.02, "Geldmarkt EUR, für Sharpe", "rf"),
        ("Angenommener Verkaufskanal",
         result["summary"].get("verkaufskanal", "cardmarket.com"),
         "Nettowerte und Break-even beziehen sich auf diesen Kanal", None),
    ]
    r = 3
    _head(ws, r, ["Parameter", "Wert", "Quelle / Begründung"])
    for label, value, src, name in rows:
        r += 1
        ws.cell(row=r, column=1, value=label).font = INK_B
        c = ws.cell(row=r, column=2, value=value)
        c.font = INPUT
        c.fill = ASSUM_FILL
        if isinstance(value, float) and value < 1:
            c.number_format = PCT
        ws.cell(row=r, column=3, value=src or "").font = INK
        if name:
            wb.defined_names.add(DefinedName(name, attr_text=f"Annahmen!$B${r}"))
    r += 2
    ws.cell(row=r, column=1, value="Legende").font = INK_B
    for txt in ["Blaue Schrift = Eingabe/Annahme (darf geändert werden)",
                "Schwarze Schrift = Formel (nicht überschreiben)",
                "Gelbe Füllung = zentrale Annahme"]:
        r += 1
        ws.cell(row=r, column=1, value=txt).font = INK
    r += 2
    ws.cell(row=r, column=1, value="Hinweise zur Bewertung").font = INK_B
    for h in result.get("hinweise", []):
        r += 1
        c = ws.cell(row=r, column=1, value="• " + h)
        c.font = INK
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 28
    _widths(ws, [38, 18, 60, 14, 14, 14])

    # ---------------------------------------------------------- Positionen
    ws = wb.create_sheet("Positionen")
    ws["A1"] = "Positionen – Einstand, Marktwert, Ergebnis"
    ws["A1"].font = TITLE
    header = ["Artikel", "Menge", "Einstand (€)", "Einstand/Stück (€)",
              "Marktpreis/Stück (€)", "Marktwert (€)", "P&L brutto (€)",
              "P&L brutto (%)", "Netto nach Gebühren (€)", "P&L netto (€)",
              "Break-even/Stück (€)", "Quelle", "Proxy?", "Zuordnung (EN)",
              "Set", "Faktor", "Plattform", "Erster Kauf"]
    _head(ws, 3, header)
    row = 3
    for p in result["positionen"]:
        row += 1
        ws.cell(row=row, column=1, value=p["artikel"]).font = INK
        ws.cell(row=row, column=2, value=p["menge"]).font = INK
        ws.cell(row=row, column=3, value=p["kosten_eur"]).font = INK
        ws.cell(row=row, column=4, value=f"=IFERROR(C{row}/B{row},0)").font = INK
        ws.cell(row=row, column=5, value=p["preis_stueck_eur"]).font = INPUT
        # Leere Marktpreis-Zelle muss sich durch die ganze Zeile fortsetzen:
        # ein Nullwert würde in den Summen als "Marktwert 0" mitgezählt und
        # den Netto-P&L um die Anschaffungskosten der unbewerteten Position
        # verfälschen.
        ws.cell(row=row, column=6,
                value=f"=IF(E{row}=\"\",\"\",E{row}*B{row})").font = INK
        ws.cell(row=row, column=7,
                value=f"=IF(F{row}=\"\",\"\",F{row}-C{row})").font = INK
        ws.cell(row=row, column=8,
                value=f"=IF(F{row}=\"\",\"\",IFERROR(F{row}/C{row}-1,\"\"))").font = INK
        # Netto nach Gebühren des angenommenen VERKAUFSkanals (nicht des
        # Kaufkanals – verkauft wird auf Cardmarket, unabhängig davon, wo
        # gekauft wurde). Der Kanal steht in 'Annahmen' und ist änderbar.
        kanal = (p.get("verkaufskanal") or "cardmarket.com").lower()
        fee = ("gebuehr_ebay" if kanal.startswith("ebay")
               else "gebuehr_amazon" if kanal.startswith("amazon")
               else "gebuehr_cm")
        ws.cell(row=row, column=9,
                value=f"=IF(F{row}=\"\",\"\",F{row}*(1-{fee}))").font = INK
        ws.cell(row=row, column=10,
                value=f"=IF(I{row}=\"\",\"\",I{row}-C{row})").font = INK
        ws.cell(row=row, column=11,
                value=f"=IFERROR(D{row}/(1-{fee}),\"\")").font = INK
        ws.cell(row=row, column=12, value=p["quelle"]).font = INK
        ws.cell(row=row, column=13,
                value=("ja" if p["is_proxy"] else ("nein" if p["is_proxy"] == 0
                                                   else "–"))).font = INK
        ws.cell(row=row, column=14, value=p.get("product_name") or "").font = INK
        ws.cell(row=row, column=15, value=p.get("group_name") or "").font = INK
        ws.cell(row=row, column=16, value=p.get("pack_factor")).font = INK
        ws.cell(row=row, column=17, value=p["plattform"]).font = INK
        ws.cell(row=row, column=18, value=p["erster_kauf"]).font = INK
        for col in (3, 4, 5, 6, 7, 9, 10, 11):
            ws.cell(row=row, column=col).number_format = EUR
        ws.cell(row=row, column=8).number_format = PCT
        for col in range(1, len(header) + 1):
            ws.cell(row=row, column=col).border = BORDER

    total = row + 1
    ws.cell(row=total, column=1, value="Summe").font = INK_B
    for col, letter in ((2, "B"), (3, "C"), (6, "F"), (7, "G"), (9, "I"), (10, "J")):
        c = ws.cell(row=total, column=col,
                    value=f"=SUM({letter}4:{letter}{row})")
        c.font = INK_B
        c.number_format = NUM if col == 2 else EUR
    c = ws.cell(row=total, column=8, value=f"=IFERROR(F{total}/C{total}-1,\"\")")
    c.font = INK_B
    c.number_format = PCT
    _widths(ws, [42, 8, 14, 16, 17, 14, 14, 12, 18, 14, 17, 26, 8, 40, 26, 8, 16, 13])

    # ------------------------------------------------------------- Verlauf
    ws = wb.create_sheet("Verlauf")
    ws["A1"] = "Portfolioverlauf (Bewertung vs. Einstand) und zeitgewichtete Rendite"
    ws["A1"].font = TITLE
    _head(ws, 3, ["Datum", "Marktwert (€)", "Einstand (€)", "Buchgewinn (€)",
                  "TWR-Index (Basis 100)"])
    twr = dict(result.get("twr") or [])
    row = 3
    for d, wert, kosten in result.get("serie", []):
        row += 1
        ws.cell(row=row, column=1, value=d).font = INK
        ws.cell(row=row, column=2, value=wert).font = INK
        ws.cell(row=row, column=3, value=kosten).font = INK
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").font = INK
        v = twr.get(d)
        if v is not None:
            ws.cell(row=row, column=5, value=v).font = INK
        for col in (2, 3, 4):
            ws.cell(row=row, column=col).number_format = EUR
    _widths(ws, [14, 16, 16, 16, 20])

    # ----------------------------------------------------------- Benchmark
    bench = result.get("benchmark") or {}
    if bench.get("series"):
        ws = wb.create_sheet("Benchmark")
        ws["A1"] = "Vergleich (alle Reihen auf 100 normiert, gemeinsamer Start)"
        ws["A1"].font = TITLE
        names = list(bench["series"].keys())
        _head(ws, 3, ["Datum"] + names)
        alldates = sorted({d for s in bench["series"].values() for d, _v in s})
        maps = {n: dict(bench["series"][n]) for n in names}
        row = 3
        for d in alldates:
            row += 1
            ws.cell(row=row, column=1, value=d).font = INK
            for i, n in enumerate(names, 2):
                v = maps[n].get(d)
                if v is not None:
                    ws.cell(row=row, column=i, value=v).font = INK
        row += 2
        ws.cell(row=row, column=1, value="Performance (%)").font = INK_B
        for i, n in enumerate(names, 2):
            c = ws.cell(row=row, column=i, value=bench["performance_pct"].get(n))
            c.font = INK_B
        _widths(ws, [14] + [14] * len(names))

    # -------------------------------------------------------- Nebenkosten
    if result.get("nebenkosten"):
        ws = wb.create_sheet("Nebenkosten")
        ws["A1"] = "Nebenkosten (nicht in den Positionskosten enthalten)"
        ws["A1"].font = TITLE
        _head(ws, 3, ["Name", "Typ", "Wert (€)", "Anzahl", "Preis/Stück (€)"])
        row = 3
        for e in result["nebenkosten"]:
            row += 1
            ws.cell(row=row, column=1, value=e["name"]).font = INK
            ws.cell(row=row, column=2, value=e["typ"]).font = INK
            ws.cell(row=row, column=3, value=e["wert_eur"]).font = INK
            ws.cell(row=row, column=4, value=e["anzahl"]).font = INK
            ws.cell(row=row, column=5,
                    value=f"=IFERROR(C{row}/D{row},0)").font = INK
            ws.cell(row=row, column=3).number_format = EUR
            ws.cell(row=row, column=5).number_format = EUR
        row += 1
        ws.cell(row=row, column=1, value="Summe").font = INK_B
        c = ws.cell(row=row, column=3, value=f"=SUM(C4:C{row - 1})")
        c.font = INK_B
        c.number_format = EUR
        _widths(ws, [34, 18, 14, 10, 16])

    # ------------------------------------------------------ Nicht bewertet
    if result.get("nicht_bewertet"):
        ws = wb.create_sheet("Offen")
        ws["A1"] = "Positionen ohne Bewertung – Zuordnung in portfolio_map.csv ergänzen"
        ws["A1"].font = TITLE
        _head(ws, 3, ["Artikel", "Grund", "Einstand (€)"])
        row = 3
        for u in result["nicht_bewertet"]:
            row += 1
            ws.cell(row=row, column=1, value=u["artikel"]).font = INK
            ws.cell(row=row, column=2, value=u["grund"]).font = INK
            c = ws.cell(row=row, column=3, value=u["kosten_eur"])
            c.font = INK
            c.number_format = EUR
        _widths(ws, [46, 28, 14])

    wb.save(path)
