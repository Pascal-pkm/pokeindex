# -*- coding: utf-8 -*-
"""Portfolio: reale Käufe bewerten, P&L und zeitgewichtete Rendite rechnen.

Das war die größte Lücke des Projekts: Order book.xlsx (die echten Käufe) und
die tägliche Preismaschinerie hatten keine Verbindung. Ohne sie gibt es keine
Rendite, kein P&L und keinen Vergleich gegen Benchmarks – also keine
Anlageanalyse, nur Marktbeobachtung.

Ablauf
------
1. `load_orders`      Excel einlesen, normalisieren, validieren (Duplikate,
                      Preis-/Mengen-Plausibilität, Summenkonsistenz).
2. `suggest_mapping`  jeden deutschen Artikelnamen einem englischen
                      TCGplayer-Produkt vorschlagen (Set-/Typ-/Token-Match),
                      Ergebnis als editierbare CSV.
3. `load_mapping`     bestätigte Zuordnung lesen (Nutzerkorrekturen gewinnen).
4. `valuate`          Tagesbewertung je Lot, Portfolio-Zeitreihe, P&L brutto
                      und netto nach Gebühren, TWR, Benchmarkvergleich.

Bewertungshierarchie je Position (transparent gekennzeichnet):
  1. manuelle Beobachtung (Cardmarket-EUR aus portfolio_prices_manual.csv)
  2. TCGplayer-Marktpreis des zugeordneten englischen Produkts, USD->EUR über
     EZB-Kurs, optional × pack_factor  ->  is_proxy = 1
  3. keine Bewertung (Position wird ausgewiesen, aber nicht mitgezählt)

Grenzen, die bewusst offengelegt werden:
  * Deutsche Auflagen sind ein anderer Markt als englische; der Proxy kann
    systematisch abweichen (Größe des Displays, Sprachprämie).
  * TCGplayer Market Price ist ein Verkaufspreis-Schätzer, kein Gebot.
  * Erlöse sind Brutto; `netto` zieht die Plattformgebühren ab (fees.py).
"""
from __future__ import annotations

import csv
import datetime as dt
import os
from dataclasses import dataclass, field

from . import fees as fee_mod
from . import names as nm
from .atomicio import write_csv

MAP_FIELDS = ["artikel", "product_id", "product_name", "group_name",
              "kategorie", "pack_factor", "sprache", "score", "status",
              "hinweis"]
MANUAL_FIELDS = ["artikel", "datum", "preis_eur", "quelle"]

# Der Bestand ist deutschsprachig; ohne Gegenbeleg gilt EUR/Cardmarket-Kontext.
DEFAULT_LANGUAGE = "DE"
MANUAL_PRICE_MAX_AGE_DAYS = 45


# --------------------------------------------------------------------- Käufe
@dataclass
class Lot:
    artikel: str
    menge: int
    gesamtpreis: float          # EUR inkl. Versand, wie gezahlt
    stueckpreis: float          # EUR je Stück
    kaufdatum: str              # ISO
    plattform: str

    @property
    def kosten(self) -> float:
        return self.gesamtpreis


def _to_iso(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def load_orders(path: str, report=None) -> list:
    """Order book.xlsx -> [Lot]. Validiert und meldet Auffälligkeiten."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [nm.norm(str(c or "")) for c in rows[0]]

    def col(*cands, default=None):
        for cand in cands:
            c = nm.norm(cand)
            for i, h in enumerate(header):
                if h == c or (c and c in h):
                    return i
        return default

    i_art = col("artikel", "produkt", "name", default=0)
    i_qty = col("menge", "anzahl", default=1)
    i_tot = col("preis inkl versand", "gesamtpreis", "preis", default=2)
    i_unit = col("preis pro stuck", "stuckpreis", default=3)
    i_date = col("kaufdatum", "datum", default=4)
    i_plat = col("plattform", "quelle", default=5)

    lots, seen = [], {}
    for r in rows[1:]:
        if not r or r[i_art] in (None, ""):
            continue
        artikel = str(r[i_art]).strip()
        try:
            menge = int(r[i_qty] or 1)
        except (TypeError, ValueError):
            menge = 1
        try:
            total = float(r[i_tot] or 0)
        except (TypeError, ValueError):
            total = 0.0
        try:
            unit = float(r[i_unit] or 0) or (total / menge if menge else 0.0)
        except (TypeError, ValueError):
            unit = total / menge if menge else 0.0
        datum = _to_iso(r[i_date]) or dt.date.today().isoformat()
        plat = str(r[i_plat] or "").strip().lower()

        if report is not None:
            if menge <= 0:
                report.warn(f"Order '{artikel}' ({datum}): Menge {menge}")
            if total <= 0:
                report.warn(f"Order '{artikel}' ({datum}): Gesamtpreis {total}")
            elif menge and abs(unit * menge - total) > max(0.05, 0.01 * total):
                report.warn(f"Order '{artikel}' ({datum}): Stückpreis × Menge "
                            f"({unit * menge:.2f}) ≠ Gesamtpreis ({total:.2f})")
            key = (nm.norm(artikel), datum, round(total, 2), menge)
            if key in seen:
                report.warn(f"Order '{artikel}' ({datum}, {total:.2f} EUR) "
                            f"erscheint mehrfach – Doppelerfassung?")
            seen[key] = True

        lots.append(Lot(artikel, menge, total, unit, datum, plat))
    lots.sort(key=lambda x: (x.kaufdatum, x.artikel))
    return lots


def load_expenditures(path: str) -> list:
    """Other Expenditures.xlsx -> [(name, typ, wert, anzahl)] (Nebenkosten)."""
    if not os.path.isfile(path):
        return []
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r or r[0] in (None, ""):
            continue
        try:
            wert = float(r[2] or 0)
        except (TypeError, ValueError):
            wert = 0.0
        try:
            anzahl = int(r[3] or 1)
        except (TypeError, ValueError):
            anzahl = 1
        out.append((str(r[0]).strip(), str(r[1] or "").strip(), wert, anzahl))
    return out


# ------------------------------------------------------------------ Matching
# Tokens, die eine TEURERE oder abweichende Sonderausgabe kennzeichnen. Steht
# ein solches Token im Kandidatennamen, aber nicht im Artikelnamen, ist die
# Zuordnung falsch: "Top Trainer Box" ist NICHT die "Pokemon Center Elite
# Trainer Box (Exclusive)" und ein "Display" ist NICHT ein "Booster Box Case".
DISTINGUISHING = {
    "case", "pokemon", "center", "exclusive", "special", "signed", "promo",
    "espanol", "japanese", "korean", "chinese", "italian", "french", "german",
    "spanish", "portuguese", "staff", "prerelease", "collector", "collectors",
    "ultra", "super", "hyper", "mini", "jumbo", "double", "triple",
}
# Generische Typwörter, die für die Unterscheidung nichts beitragen.
GENERIC = {"top", "trainer", "box", "display", "booster", "kollektion", "er",
           "pack", "blister", "bundle", "elite", "collection", "premium",
           "tin", "deck", "the", "and", "of", "pokemon", "tcg", "set"}


def suggest_mapping(articles, products: dict) -> list:
    """Zuordnungsvorschläge deutscher Artikel -> englische Produkte.

    products: {product_id: stammdatensatz aus products.csv.gz}
    Rückgabe: Liste von Mapping-Dicts (MAP_FIELDS) mit Score 0..1.

    Scoring: Typtreffer (stärkstes Signal) + Kernbegriffe + übersetzte
    Zusatztokens MINUS Abzug für unterscheidende Zusätze, die nur der Kandidat
    trägt. Ohne diesen Abzug gewinnen systematisch die Sonderausgaben, weil
    ihre längeren Namen mehr Trefferchancen haben.
    """
    by_group = {}
    for pid, p in products.items():
        if str(p.get("is_sealed")) != "1":
            continue
        by_group.setdefault(p.get("group_name", ""), []).append((pid, p))

    out = []
    for artikel in sorted(set(articles)):
        de_set, en_group = nm.detect_set(artikel)
        cat, en_terms = nm.detect_type(artikel)
        packs = nm.detect_pack_count(artikel)
        extras = nm.translate_extras(artikel)
        cands = by_group.get(en_group, []) if en_group else []
        art_tokens = extras | nm.tokens(artikel)

        best = (0.0, None, None)
        second = (0.0, None, None)
        for pid, p in cands:
            pname = p.get("name") or ""
            pt = nm.tokens(pname)
            score = 0.0
            # Typtreffer ist das stärkste Signal
            if cat and p.get("sealed_cat") == cat:
                score += 0.45
            elif cat and p.get("sealed_cat") and p.get("sealed_cat") != cat:
                score -= 0.35          # falscher Produkttyp (z. B. Case)
            if en_terms and any(t in nm.norm(pname) for t in en_terms):
                score += 0.25
            # Restliche Tokens (Pokémon-Namen, Varianten wie "Pokemon Center")
            set_tokens = nm.tokens(de_set or "") | nm.tokens(en_group or "")
            rel = {t for t in extras if t not in set_tokens and t not in GENERIC}
            if rel:
                score += 0.30 * (len(rel & pt) / len(rel))
            elif cat and p.get("sealed_cat") == cat:
                score += 0.15          # generisches Produkt, kein Zusatz nötig
            # Abzug: Sonderausgabe-Marker nur im Kandidaten
            extra_marks = {t for t in (pt - art_tokens) - set_tokens
                           if t in DISTINGUISHING}
            score -= 0.30 * len(extra_marks)
            if score > best[0]:
                second = best
                best = (score, pid, p)
            elif score > second[0]:
                second = (score, pid, p)

        score, pid, p = best
        hinweise = []
        if not en_group:
            hinweise.append("Set nicht erkannt")
        if not cat:
            hinweise.append("Produkttyp nicht erkannt")
        # Mehrdeutigkeit: liegen zwei Kandidaten dicht beieinander, ist die Wahl
        # nicht datengestützt (typisch bei Sets mit zwei ETB-Varianten, z. B.
        # Ice Rider / Shadow Rider Calyrex).
        ambiguous = (second[1] is not None and score - second[0] <= 0.05)
        if ambiguous:
            hinweise.append(f"mehrdeutig – Alternative: {second[2].get('name')} "
                            f"(ID {second[1]})")
        pack_factor = 1.0
        # Deutsche Displays haben je Ära 18 oder 36 Päckchen; das englische
        # Vergleichsprodukt ist üblicherweise eine 36er-Box. Ohne Korrektur
        # wäre ein 18er-Display doppelt so hoch bewertet.
        if packs and packs < 30:
                pack_factor = round(packs / 36.0, 4)
                hinweise.append(f"{packs}er-Display gegen 36er-Box skaliert "
                                f"(Faktor {pack_factor}) – bitte prüfen")
        if pid is None:
            status = "offen"
        elif ambiguous or score < 0.55:
            status = "unsicher"
        else:
            status = "auto"
        out.append({
            "artikel": artikel,
            "product_id": pid or "",
            "product_name": (p.get("name") if p else "") or "",
            "group_name": (p.get("group_name") if p else en_group) or "",
            "kategorie": cat or "",
            "pack_factor": pack_factor,
            "sprache": DEFAULT_LANGUAGE,
            "score": round(score, 3),
            "status": status,
            "hinweis": "; ".join(hinweise),
        })
    return out


def write_mapping(path: str, rows) -> None:
    """Zuordnungsdatei schreiben – mit BOM, damit Excel die Umlaute richtig
    anzeigt (die Datei ist zum Bearbeiten durch den Nutzer gedacht)."""
    write_csv(path, MAP_FIELDS,
              [[r.get(k, "") for k in MAP_FIELDS] for r in rows], bom=True)


def load_mapping(path: str) -> dict:
    """{normalisierter artikel: mapping-dict}. Nutzerkorrekturen gewinnen."""
    if not os.path.isfile(path):
        return {}
    out = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            key = nm.norm(row.get("artikel", ""))
            if not key:
                continue
            try:
                row["pack_factor"] = float(row.get("pack_factor") or 1.0)
            except ValueError:
                row["pack_factor"] = 1.0
            pid = (row.get("product_id") or "").strip()
            row["product_id"] = int(pid) if pid.isdigit() else None
            out[key] = row
    return out


def merge_mapping(existing: dict, suggestions: list) -> list:
    """Bestehende (ggf. händisch korrigierte) Zuordnung mit neuen Vorschlägen
    zusammenführen: bestätigte Zeilen bleiben unverändert."""
    out = []
    for s in suggestions:
        key = nm.norm(s["artikel"])
        cur = existing.get(key)
        if cur and (cur.get("status") in {"bestaetigt", "bestätigt", "manuell"}
                    or cur.get("product_id")):
            merged = dict(s)
            merged.update({k: cur.get(k, s.get(k)) for k in MAP_FIELDS
                           if cur.get(k) not in (None, "")})
            out.append(merged)
        else:
            out.append(s)
    # Zeilen, die nur in der Datei stehen (händisch ergänzt), erhalten
    known = {nm.norm(r["artikel"]) for r in out}
    for key, cur in existing.items():
        if key not in known:
            out.append(cur)
    return out


def refresh_labels(rows, products: dict) -> list:
    """Produktname/Set aus der product_id neu setzen.

    Korrigiert der Nutzer in der CSV nur die ID (der häufige Fall bei
    mehrdeutigen Varianten), stünde sonst weiter der alte Name daneben – die
    Bewertung wäre richtig, die Anzeige irreführend.
    """
    for r in rows:
        pid = r.get("product_id")
        try:
            pid = int(pid) if pid not in (None, "") else None
        except (TypeError, ValueError):
            pid = None
        if pid is None:
            continue
        p = products.get(pid)
        if not p:
            r["hinweis"] = ((r.get("hinweis") or "") +
                            f"; product_id {pid} ist unbekannt").lstrip("; ")
            continue
        r["product_id"] = pid
        r["product_name"] = p.get("name") or ""
        r["group_name"] = p.get("group_name") or ""
        r["kategorie"] = p.get("sealed_cat") or r.get("kategorie") or ""
    return rows


def load_manual_prices(path: str) -> dict:
    """{normalisierter artikel: [(datum, preis_eur, quelle), ...]} sortiert."""
    out: dict = {}
    if not os.path.isfile(path):
        return out
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            key = nm.norm(row.get("artikel", ""))
            d = _to_iso(row.get("datum"))
            try:
                price = float(str(row.get("preis_eur", "")).replace(",", "."))
            except (TypeError, ValueError):
                continue
            if key and d and price > 0:
                out.setdefault(key, []).append((d, price, row.get("quelle", "")))
    for k in out:
        out[k].sort()
    return out


def _manual_on_or_before(obs, day: str):
    best = None
    for d, price, src in obs:
        if d <= day:
            best = (d, price, src)
        else:
            break
    if not best:
        return None
    age = (dt.date.fromisoformat(day) - dt.date.fromisoformat(best[0])).days
    if age > MANUAL_PRICE_MAX_AGE_DAYS:
        return None
    return best


# ---------------------------------------------------------------- Bewertung
@dataclass
class Valuation:
    positions: list = field(default_factory=list)
    series: list = field(default_factory=list)      # [[datum, wert_eur, kosten_eur]]
    twr: list = field(default_factory=list)         # [[datum, index]] Basis 100
    summary: dict = field(default_factory=dict)
    unmatched: list = field(default_factory=list)


def valuate(lots, mapping: dict, price_series: dict, rates: dict,
            manual: dict | None = None, dates: list | None = None,
            products: dict | None = None,
            sell_platform: str = "cardmarket.com") -> Valuation:
    """Portfolio bewerten.

    lots:          [Lot]
    mapping:       {norm(artikel): mapping-dict}
    price_series:  {product_id: {datum: cents_usd}}  (TCGplayer)
    rates:         EZB-Kursreihe USD->EUR
    manual:        {norm(artikel): [(datum, preis_eur, quelle)]}
    dates:         Bewertungsraster (sortierte ISO-Daten); Standard = alle
                   Tage von erstem Kauf bis letztem Preisdatum
    sell_platform: angenommener VERKAUFSkanal für Nettowerte und Break-even.

    Wichtig: Nettowerte richten sich nach dem Verkaufs-, nicht dem Kaufkanal.
    Ein bei Pokémon Center gekauftes Produkt wird nicht dort verkauft; die
    frühere Berechnung nach Kaufplattform setzte für solche Positionen 0 %
    Gebühren an und überschätzte den erzielbaren Erlös.
    """
    from . import fx
    manual = manual or {}
    products = products or {}

    all_price_dates = sorted({d for s in price_series.values() for d in s})
    if not all_price_dates:
        return Valuation(summary={"fehler": "keine Preisdaten"})
    if dates is None:
        first = min(lot.kaufdatum for lot in lots) if lots else all_price_dates[0]
        start = max(first, all_price_dates[0])
        end = all_price_dates[-1]
        d0, d1 = dt.date.fromisoformat(start), dt.date.fromisoformat(end)
        dates = [(d0 + dt.timedelta(days=i)).isoformat()
                 for i in range((d1 - d0).days + 1)]

    # Preis-Lookup mit Forward-Fill innerhalb der Reihe
    filled: dict = {}
    for pid, s in price_series.items():
        ds = sorted(s)
        cur = {}
        last = None
        j = 0
        for day in dates:
            while j < len(ds) and ds[j] <= day:
                last = s[ds[j]]
                last_day = ds[j]
                j += 1
            if last is not None and (dt.date.fromisoformat(day)
                                     - dt.date.fromisoformat(last_day)).days <= 70:
                cur[day] = last
        filled[pid] = cur

    def price_eur(artikel_key: str, m: dict | None, day: str):
        """(preis_eur, quelle, is_proxy) oder (None, grund, None)."""
        obs = manual.get(artikel_key)
        if obs:
            hit = _manual_on_or_before(obs, day)
            if hit:
                return hit[1], f"manuell ({hit[2] or 'Cardmarket'}, {hit[0]})", 0
        if not m or not m.get("product_id"):
            return None, "keine Zuordnung", None
        pid = m["product_id"]
        cents = filled.get(pid, {}).get(day)
        if cents is None:
            return None, "kein Preis am Tag", None
        usd = cents / 100.0 * float(m.get("pack_factor") or 1.0)
        eur = fx.usd_to_eur(usd, day, rates)
        if eur is None:
            return None, "kein FX-Kurs", None
        return eur, "TCGplayer EN (Proxy)", 1

    # ---- Positionen (Aggregat je Artikel) ----
    by_article: dict = {}
    for lot in lots:
        key = nm.norm(lot.artikel)
        agg = by_article.setdefault(key, {
            "artikel": lot.artikel, "menge": 0, "kosten": 0.0,
            "erster_kauf": lot.kaufdatum, "letzter_kauf": lot.kaufdatum,
            "plattformen": set(), "lots": []})
        agg["menge"] += lot.menge
        agg["kosten"] += lot.gesamtpreis
        agg["erster_kauf"] = min(agg["erster_kauf"], lot.kaufdatum)
        agg["letzter_kauf"] = max(agg["letzter_kauf"], lot.kaufdatum)
        agg["plattformen"].add(lot.plattform or "privat")
        agg["lots"].append(lot)

    today = dates[-1]
    positions, unmatched = [], []
    proxy_value = manual_value = 0.0
    for key, agg in sorted(by_article.items(), key=lambda kv: -kv[1]["kosten"]):
        m = mapping.get(key)
        px, src, is_proxy = price_eur(key, m, today)
        stueck_kosten = agg["kosten"] / max(agg["menge"], 1)
        plat = sorted(agg["plattformen"])[0]          # Kaufkanal (Dokumentation)
        wert = px * agg["menge"] if px is not None else None
        netto = (fee_mod.net_of_fees(px, sell_platform, agg["menge"])
                 if px is not None else None)
        pos = {
            "artikel": agg["artikel"],
            "menge": agg["menge"],
            "kosten_eur": round(agg["kosten"], 2),
            "kosten_stueck_eur": round(stueck_kosten, 2),
            "erster_kauf": agg["erster_kauf"], "letzter_kauf": agg["letzter_kauf"],
            "plattform": plat,
            "product_id": (m or {}).get("product_id"),
            "product_name": (m or {}).get("product_name"),
            "group_name": (m or {}).get("group_name"),
            "pack_factor": float((m or {}).get("pack_factor") or 1.0),
            "preis_stueck_eur": round(px, 2) if px is not None else None,
            "wert_eur": round(wert, 2) if wert is not None else None,
            "wert_netto_eur": round(netto, 2) if netto is not None else None,
            "quelle": src,
            "is_proxy": is_proxy,
            "pl_eur": round(wert - agg["kosten"], 2) if wert is not None else None,
            "pl_pct": (round((wert / agg["kosten"] - 1) * 100, 2)
                       if wert is not None and agg["kosten"] > 0 else None),
            "pl_netto_eur": (round(netto - agg["kosten"], 2)
                             if netto is not None else None),
            "pl_netto_pct": (round((netto / agg["kosten"] - 1) * 100, 2)
                             if netto is not None and agg["kosten"] > 0 else None),
            "breakeven_stueck_eur": round(
                fee_mod.breakeven_price(stueck_kosten, sell_platform), 2),
            "verkaufskanal": sell_platform,
        }
        positions.append(pos)
        if wert is None:
            unmatched.append({"artikel": agg["artikel"], "grund": src,
                              "kosten_eur": round(agg["kosten"], 2)})
        elif is_proxy:
            proxy_value += wert
        else:
            manual_value += wert

    # ---- Zeitreihe + zeitgewichtete Rendite (TWR) ----
    # Korrekte TWR: die Tagesrendite wird ausschließlich auf dem Bestand des
    # VORTAGS gemessen (zu den heutigen Preisen), verglichen mit dem
    # Gesamtwert des Vortags. Neu gekaufte Positionen gehen erst ab dem
    # Folgetag in die Renditebasis ein.
    #
    # Der naive Ansatz (Gesamtwert heute minus Kaufkosten von heute, geteilt
    # durch Gesamtwert gestern) ist falsch, sobald Marktwert und Kaufpreis
    # auseinanderliegen: die Differenz erscheint als Tagesrendite. Bei einem
    # Bestand, dessen Bewertung deutlich über dem Einstand liegt, führt das zu
    # absurden Renditen (im ersten Testlauf über 5.000 %).
    series, twr = [], []
    prev_total = None
    prev_day = None
    idx = 100.0
    for day in dates:
        total = 0.0
        cost = 0.0
        base_now = 0.0            # Wert des Vortagsbestands zu heutigen Preisen
        have_any = False
        for key, agg in by_article.items():
            m = mapping.get(key)
            qty_now = sum(lot.menge for lot in agg["lots"] if lot.kaufdatum <= day)
            if qty_now <= 0:
                continue
            cost += sum(lot.gesamtpreis for lot in agg["lots"] if lot.kaufdatum <= day)
            px, _src, _p = price_eur(key, m, day)
            if px is None:
                continue
            have_any = True
            total += px * qty_now
            if prev_day is not None:
                qty_prev = sum(lot.menge for lot in agg["lots"]
                               if lot.kaufdatum <= prev_day)
                base_now += px * qty_prev
        if not have_any:
            continue
        series.append([day, round(total, 2), round(cost, 2)])
        if prev_total is None:
            twr.append([day, 100.0])
        elif prev_total > 0:
            idx *= (base_now / prev_total)
            twr.append([day, round(idx, 2)])
        prev_total, prev_day = total, day

    total_cost = sum(p["kosten_eur"] for p in positions)
    total_val = sum(p["wert_eur"] or 0 for p in positions)
    total_net = sum(p["wert_netto_eur"] or 0 for p in positions)
    valued_cost = sum(p["kosten_eur"] for p in positions if p["wert_eur"] is not None)
    summary = {
        "stand": today,
        "positionen": len(positions),
        "stueckzahl": sum(p["menge"] for p in positions),
        "kosten_eur": round(total_cost, 2),
        "kosten_bewertet_eur": round(valued_cost, 2),
        "marktwert_eur": round(total_val, 2),
        "marktwert_netto_eur": round(total_net, 2),
        "pl_eur": round(total_val - valued_cost, 2),
        "pl_pct": (round((total_val / valued_cost - 1) * 100, 2)
                   if valued_cost > 0 else None),
        "pl_netto_eur": round(total_net - valued_cost, 2),
        "pl_netto_pct": (round((total_net / valued_cost - 1) * 100, 2)
                         if valued_cost > 0 else None),
        "proxy_anteil_pct": (round(100 * proxy_value / total_val, 1)
                             if total_val > 0 else None),
        "nicht_bewertet": len(unmatched),
        "nicht_bewertet_kosten_eur": round(
            sum(u["kosten_eur"] for u in unmatched), 2),
        "twr_index": twr[-1][1] if twr else None,
        "twr_pct": round(twr[-1][1] - 100, 2) if twr else None,
        "verkaufskanal": sell_platform,
        "verkaufsgebuehr_pct": round(
            100 * fee_mod.model_for(sell_platform).total_fee_pct(100.0), 2),
    }
    return Valuation(positions=positions, series=series, twr=twr,
                     summary=summary, unmatched=unmatched)


def benchmark_comparison(twr: list, benchmarks: dict) -> dict:
    """Portfolio-TWR gegen Benchmarks über den gemeinsamen Zeitraum (Basis 100)."""
    from . import risk
    if not twr:
        return {}
    start = twr[0][0]
    out = {"Portfolio": risk.rebase(twr, 100.0, start)}
    for name, series in benchmarks.items():
        reb = risk.rebase([p for p in series if p[0] >= start], 100.0, start)
        if reb:
            out[name] = reb
    perf = {}
    for name, s in out.items():
        if len(s) >= 2:
            perf[name] = round(s[-1][1] - 100, 2)
    return {"series": out, "performance_pct": perf}
